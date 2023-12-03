import numpy as np
import pandas as pd
import openpyxl as op
from datetime import date

class TransgenicMouseData():

    def __init__(self):
        self.wb = op.load_workbook('/Users/veramazeeva/Library/CloudStorage/OneDrive-Personal/Carnegie Mellon/CompBio512/project/TransgenicMouseData.xlsx')
        self.mouse2exp = {self.wb['Main experiment summary']['B'+str(6+i)].value: [self.wb['Main experiment summary']['E'+str(6+i)].value, self.wb['Main experiment summary']['I'+str(6+i)].value] for i in range(0, 58)}
        self.mouse2data = dict()
        self.untreated_mice = list()
        self.treated_mice = list()
        self.__init__mouse_data()
    
    def __init__mouse_data(self):
        mouse2cell = dict()
        mouse_ids = set(self.mouse2exp.keys())
        for mouse_exp in self.mouse2exp.values():
            sheet = self.wb['Tumor sizes_' + mouse_exp[0]]
            for row in sheet.iter_rows(min_row=8, min_col=4, max_row=270, max_col=5, values_only=False):
                for cell in row:
                    if cell.value in mouse_ids:
                        mouse2cell[cell.value] = (cell.row, cell.column)
                        break

        for mouse_id, mouse_cell in mouse2cell.items():
            if mouse_id == 'CM.41':
                continue
            
            self.mouse2data[mouse_id] = dict()
            
            r, c = mouse_cell[0] + 2, mouse_cell[1]
            date_row = r-3
            sheet = self.wb['Tumor sizes_' + self.mouse2exp[mouse_id][0]]
            
            H, V, D = [], [], []
            while sheet.cell(r,c).value != None:
           
                # extract only Tumor 1 measurements 
                h, v = [], []
                for i in range(3):
                    measure_h = str(sheet.cell(r+i, c).value)
                    measure_v = str(sheet.cell(r+3+i, c).value)
                    h.append(float(measure_h)) if '*' not in measure_h else h.append(float(measure_h[:-1]))
                    v.append(float(measure_v)) if '*' not in measure_v else v.append(float(measure_v[:-1]))
    
                d = (str(sheet.cell(date_row, c).value)).split("/")
                day, month = int(d[0] if d[0]!=0 else d[0][-1]), int(d[1] if d[1]!=0 else d[1][-1])
                D.append(date(2012, month, day))

                H.append(np.median(h))
                V.append(np.median(v))
                
                c += 1
       
            H = np.array(H)
            V = np.array(V)
            
            self.mouse2data[mouse_id]['Volume'] = np.pi/6 * np.multiply(np.multiply(H,V), (1/2)*(H+V))
            self.mouse2data[mouse_id]['Date'] = D

            if self.mouse2exp[mouse_id][1] == 'DMSO':
                self.untreated_mice.append(mouse_id)
        
            else:
                self.treated_mice.append(mouse_id)

